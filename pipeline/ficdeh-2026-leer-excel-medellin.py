# -*- coding: utf-8 -*-
"""Lee el Excel OFICIAL de programación de Medellín → sidecar de staging.

Es la primera fuente que el festival entrega en firme, y la de mayor autoridad
hasta ahora: trae por función fecha, hora, sede, categoría, título, dirección,
duración, país, año, SINOPSIS EN ESPAÑOL E INGLÉS (oficiales, no traducidas por
nosotros), póster, tráiler, temáticas y clasificación de público.

Las filas con RELLENO ROJO están dadas de baja: el festival las sacó de la
programación (aviso de Juan). El color es el único sitio donde consta —el texto
de la fila no dice nada—, así que se lee el formato y se marcan `en_app: False`.

NO se copian al sidecar los enlaces de screener ni sus contraseñas (columnas
LINK/CONTRASEÑA): son material de visionado privado del festival y no tienen
por qué viajar a un JSON que se publica.

Solo cubre MEDELLÍN. Las otras diez ciudades siguen dependiendo del sitio web.
"""
import json, os, re, datetime
import openpyxl

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.expanduser('~/Desktop/Programación _ MEDELLÍN.xlsx')
OUT = f'{REPO}/festivals/staging/ficdeh-2026-medellin-oficial.json'

COLS = {'fecha': 1, 'hora': 2, 'sede': 3, 'categoria': 4, 'titulo': 5, 'director': 6,
        'duracion': 7, 'pais': 8, 'anio': 9, 'synopsis': 10, 'synopsis_en': 11,
        'perfil_director': 12, 'redes_pelicula': 13, 'redes_director': 14,
        'poster_url': 15, 'trailer': 16, 'tematica': 17, 'tematica_2': 18,
        'kit_prensa': 19, 'clasificacion': 22}


def limpio(v):
    if v is None:
        return ''
    if isinstance(v, datetime.datetime):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return re.sub(r'\s+', ' ', str(v)).strip()


def hora24(h):
    m = re.match(r'(\d{1,2}):(\d{2})\s*([ap])\.?\s*m', h.strip(), re.I)
    if not m:
        return h.strip()
    hh, mm, ap = int(m.group(1)), m.group(2), m.group(3).lower()
    if ap == 'p' and hh != 12:
        hh += 12
    if ap == 'a' and hh == 12:
        hh = 0
    return f'{hh:02d}:{mm}'


def main():
    ws = openpyxl.load_workbook(XLSX, data_only=True)['FICDEH']
    fmt = openpyxl.load_workbook(XLSX)['FICDEH']   # data_only pierde nada de estilo, pero se lee aparte por claridad

    def anulada(r):
        """Fila dada de baja: relleno rojo. Se mira la fila entera —no una celda
        suelta— porque el festival tacha la función completa."""
        rojas = sum(1 for c in range(1, 23)
                    if (fmt.cell(r, c).fill and fmt.cell(r, c).fill.fgColor
                        and fmt.cell(r, c).fill.fgColor.rgb == 'FFFF0000'))
        return rojas >= 3
    funcs = []
    for r in range(4, ws.max_row + 1):
        get = lambda k: limpio(ws.cell(r, COLS[k]).value)
        if not get('titulo') or not get('fecha'):
            continue
        f = {k: get(k) for k in COLS}
        f['hora'] = hora24(f['hora'])
        # el país llega con saltos de línea, barras y guiones sueltos
        f['pais'] = re.sub(r'\s*[/\-]\s*|\s{2,}', ', ', f['pais']).strip(' ,')
        f['duracion'] = f"{f['duracion']} min" if f['duracion'] else ''
        f['ciudad'] = 'Medellín'
        f['en_app'] = not anulada(r)
        if not f['en_app']:
            f['_baja'] = 'marcada en rojo en el Excel: fuera de programación'
        funcs.append(f)

    json.dump({'_provenance': {
        'fuente': 'Excel oficial «Programación _ MEDELLÍN.xlsx», entregado por el festival',
        'recibido': '2026-08-06',
        'alcance': 'SOLO Medellín. Las otras 10 ciudades siguen saliendo del sitio web.',
        'autoridad': 'La más alta hasta ahora: lo entrega el festival, no lo publica. '
                     'Sus sinopsis en inglés son OFICIALES y ganan a nuestras traducciones.',
        'no_incluido': 'Enlaces de screener y contraseñas (columnas LINK/CONTRASEÑA): '
                       'material de visionado privado, no va a un JSON publicable.',
    }, 'funciones': funcs}, open(OUT, 'w', encoding='utf-8'), ensure_ascii=False, indent=1)

    obras = {f['titulo'] for f in funcs}
    print(f'{len(funcs)} funciones · {len(obras)} obras únicas')
    baja = [f for f in funcs if not f['en_app']]
    print(f'DADAS DE BAJA (rojo): {len(baja)}')
    for f in baja:
        print(f"   ✗ {f['fecha']} {f['hora']} {f['sede'][:28]:29} {f['titulo'][:34]}")
    print(f'con sinopsis EN oficial: {sum(1 for f in funcs if f["synopsis_en"])}')
    print(f'con póster: {sum(1 for f in funcs if f["poster_url"])}')
    print(f'sedes: {sorted({f["sede"] for f in funcs})}')
    print(f'días: {sorted({f["fecha"] for f in funcs})}')


if __name__ == '__main__':
    main()
